import asyncio
import json
from typing import List, Dict, Any, Callable, Optional, Union
import pandas as pd
import io
from groq import AsyncGroq
import streamlit as st
from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception_type
import aiohttp
import logging
from dataclasses import dataclass
from pydantic import BaseModel, Field
import plotly.express as px
import base64
from openpyxl import Workbook
import numpy as np

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class AnalysisRequest(BaseModel):
    text: str
    instruction: str
    model: str = Field(default="llama3-8b-8192")
    temperature: float = Field(default=0.2, ge=0, le=1)
    max_tokens: int = Field(default=1000, ge=1)


class Usage(BaseModel):
    completion_tokens: int
    prompt_tokens: int
    total_tokens: int
    completion_time: float
    prompt_time: float
    queue_time: float
    total_time: float


class AnalysisResponse(BaseModel):
    content: Dict[str, Any]
    model: str
    usage: Usage


@dataclass
class AnalysisResult:
    success: bool
    data: Optional[AnalysisResponse] = None
    error: Optional[str] = None


class GroqParser:
    def __init__(self, api_key: str):
        self.client = AsyncGroq(api_key=api_key)
        self.session = aiohttp.ClientSession()

    async def __aenter__(self):
        return self

    async def __aexit__(self, exc_type, exc, tb):
        await self.session.close()

    @retry(
        stop=stop_after_attempt(3),
        wait=wait_exponential(multiplier=1, min=4, max=10),
        retry=retry_if_exception_type((aiohttp.ClientError, asyncio.TimeoutError))
    )
    async def analyze_text(self, request: AnalysisRequest) -> AnalysisResult:
        try:
            prompt = self._get_prompt(request.text, request.instruction)

            chat_completion = await self.client.chat.completions.create(
                messages=[
                    {
                        "role": "system",
                        "content": "You are an AI assistant specialized in analyzing and extracting information from text based on specific instructions. Always provide your response in a structured JSON format."
                    },
                    {
                        "role": "user",
                        "content": prompt
                    }
                ],
                model=request.model,
                temperature=request.temperature,
                max_tokens=request.max_tokens,
            )

            parsed_content = self._parse_response(chat_completion.choices[0].message.content)

            response = AnalysisResponse(
                content=parsed_content,
                model=chat_completion.model,
                usage=Usage(**chat_completion.usage.dict())
            )

            return AnalysisResult(success=True, data=response)
        except Exception as e:
            logger.error(f"Error in analyze_text: {str(e)}")
            return AnalysisResult(success=False, error=str(e))

    @staticmethod
    def _get_prompt(text: str, instruction: str) -> str:
        return f"""Analyze the following text and {instruction}. 
        Provide your response in a clear, structured JSON format.
        Ensure all keys in the JSON are strings and all values are either strings, numbers, booleans, or arrays of these types.

        Text to analyze:
        {text}
        """

    @staticmethod
    def _parse_response(response: str) -> Dict[str, Any]:
        try:
            return json.loads(response)
        except json.JSONDecodeError:
            # If JSON parsing fails, attempt to extract JSON-like structure
            import re
            json_like = re.search(r'\{.*}', response, re.DOTALL)
            if json_like:
                try:
                    return json.loads(json_like.group())
                except json.JSONDecodeError:
                    pass
            return {"raw_response": response}


async def async_groq_parser(
        data_bits: List[str],
        instruction: str,
        progress_callback: Optional[Callable[[int, str], None]] = None,
        batch_size: int = 5
) -> List[Dict[str, Any]]:
    async with GroqParser(st.secrets["GROQ_API_KEY"]) as parser:
        async def process_batch(batch: List[str], start_index: int) -> tuple[Any]:
            tasks = [
                parser.analyze_text(AnalysisRequest(text=bit, instruction=instruction))
                for bit in batch
            ]
            results = await asyncio.gather(*tasks)
            if progress_callback:
                progress = int((start_index + len(batch)) / len(data_bits) * 100)
                progress_callback(progress, f"Analyzed {start_index + len(batch)} of {len(data_bits)} bits")
            return results

        all_results = []
        for i in range(0, len(data_bits), batch_size):
            batch = data_bits[i:i + batch_size]
            batch_results = await process_batch(batch, i)
            all_results.extend(batch_results)

        if progress_callback:
            progress_callback(100, "Analysis complete!")

        return [
            result.data.dict() if result.success else {"error": result.error}
            for result in all_results
        ]


def groq_parser(
        data_bits: List[str],
        instruction: str,
        progress_callback: Optional[Callable[[int, str], None]] = None
) -> List[Dict[str, Any]]:
    try:
        results = asyncio.run(async_groq_parser(data_bits, instruction, progress_callback))

        # Check if the instruction contains a visualization request
        viz_type = None
        if "table" in instruction.lower():
            viz_type = "table"
        elif "graph" in instruction.lower() or "chart" in instruction.lower():
            viz_type = "graph"

        # Automatically create visualization
        display_visualization(results, viz_type)

        return results
    except Exception as e:
        error_message = f"An error occurred during parsing: {str(e)}"
        logger.error(error_message, exc_info=True)
        st.error(error_message)
        return [{"error": error_message}]



def create_visualization(data: Union[list, dict], viz_type: str = None):
    if isinstance(data, list) and len(data) > 0:
        # Flatten the data structure if it contains 'data' and 'content' keys
        flattened_data = []
        for item in data:
            if isinstance(item, dict) and 'data' in item and 'content' in item['data']:
                flattened_data.append(item['data']['content'])
            else:
                flattened_data.append(item)
        df = pd.DataFrame(flattened_data)
    elif isinstance(data, dict):
        if 'data' in data and 'content' in data['data']:
            df = pd.DataFrame([data['data']['content']])
        else:
            df = pd.DataFrame([data])
    else:
        return None, 'none'

    numeric_columns = df.select_dtypes(include=['number']).columns
    categorical_columns = df.select_dtypes(include=['object']).columns

    if viz_type == 'graph' or (viz_type is None and len(numeric_columns) > 0):
        if len(numeric_columns) >= 2:
            fig = px.scatter(df, x=numeric_columns[0], y=numeric_columns[1],
                             title=f"{numeric_columns[1]} vs {numeric_columns[0]}")
            fig.update_traces(marker=dict(size=10))
        elif len(numeric_columns) == 1:
            fig = px.line(df, y=numeric_columns[0], title=f"Trend of {numeric_columns[0]}")
        elif len(categorical_columns) >= 4:
            column_to_plot = categorical_columns[0]
            fig = px.bar(df[column_to_plot].value_counts(), title=f"Distribution of {column_to_plot}")
            fig.update_layout(xaxis_title=column_to_plot, yaxis_title="Count")
        else:
            return df, 'table'

        fig.update_layout(
            template="plotly_dark",
            title_font_size=24,
            legend_title_font_size=14,
            legend_font_size=12,
            hoverlabel=dict(bgcolor="white", font_size=14),
            hovermode="closest"
        )
        return fig, 'graph'

    return df, 'table'


def display_visualization(result: Union[list, dict], _: Any = None):  # Added optional second parameter
    global content_df
    if 'result_data' in st.session_state:
        del st.session_state['result_data']
    if 'content_df' in st.session_state:
        del st.session_state['content_df']

    st.session_state.result_data = result
    print("Result Data:")
    print(result)

    # Convert to DataFrame
    if isinstance(st.session_state.result_data, dict):
        df = pd.DataFrame([st.session_state.result_data])
    elif isinstance(st.session_state.result_data, list):
        df = pd.DataFrame(st.session_state.result_data)
    else:
        df = pd.DataFrame([st.session_state.result_data])
    print("DataFrame (df):")
    print(df)

    # Parse 'content' column if it's a string or dict
    if 'content' in df.columns:
        content_data = []
        for _, row in df.iterrows():
            content = row['content']
            if isinstance(content, str):
                try:
                    content = json.loads(content)
                except json.JSONDecodeError:
                    pass
            if isinstance(content, dict):
                content_data.append(content)
            elif isinstance(content, list):
                content_data.extend(content)

        if content_data:
            content_df = pd.json_normalize(content_data)
            st.session_state.content_df = content_df
            print("Content DataFrame (content_df):")
            print(content_df)
        st.subheader("Scraped Content Visualization")

        # Determine the best visualization based on the data
        numeric_cols = content_df.select_dtypes(include=[np.number]).columns
        categorical_cols = content_df.select_dtypes(include=['object']).columns

        if len(numeric_cols) >= 2:
            # Create a scatter plot of the first two numeric columns
            fig = px.scatter(content_df, x=numeric_cols[0], y=numeric_cols[1],
                             title=f"{numeric_cols[1]} vs {numeric_cols[0]}")
            st.plotly_chart(fig, use_container_width=True)

        elif len(numeric_cols) == 1:
            # Create a histogram of the single numeric column
            fig = px.histogram(content_df, x=numeric_cols[0],
                               title=f"Distribution of {numeric_cols[0]}")
            st.plotly_chart(fig, use_container_width=True)

        elif len(categorical_cols) > 0:
            # Create a bar chart of the first categorical column
            fig = px.bar(content_df[categorical_cols[0]].value_counts(),
                         title=f"Counts of {categorical_cols[0]}")
            st.plotly_chart(fig, use_container_width=True)

        # Always display the data in a table
        st.subheader("Scraped Data Table")
        st.dataframe(content_df)

        # Download options
        st.subheader("Download Options")
        col1, col2 = st.columns(2)
        with col1:
            csv = content_df.to_csv(index=False)
            st.download_button(
                label="Download as CSV",
                data=csv,
                file_name="scraped_data.csv",
                mime="text/csv",
            )
        with col2:
            # Create Excel file
            wb = Workbook()
            ws = wb.active
            ws.title = "Scraped Data"

            # Write data
            for r in dataframe_to_rows(content_df, index=False, header=True):
                ws.append(r)

            # Save to BytesIO
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            # Create download link
            b64 = base64.b64encode(excel_buffer.getvalue()).decode()
            href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="scraped_data.xlsx">Download Excel file</a>'
            st.markdown(href, unsafe_allow_html=True)
    else:
        st.warning("No structured content data available for visualization.")

    if 'result_data' in st.session_state and 'content_df' in st.session_state:
        result_data = st.session_state.result_data
        if isinstance(result_data, dict):
            result_df = pd.DataFrame([result_data])
        elif isinstance(result_data, list):
            result_df = pd.DataFrame(result_data)
        else:
            result_df = result_data

def dataframe_to_rows(df, index=False, header=True):
    rows = []
    if header:
        rows.append(df.columns.tolist())
    for row in df.itertuples(index=index):
        rows.append([str(x) for x in row[1:]])
    return rows


def display_debug_info(df, content_df):
    st.write("Debug Info:")
    st.write("AI Metrics DataFrame:")
    st.write(f"Shape: {df.shape}")
    st.write(f"Columns: {df.columns.tolist()}")
    st.write(f"Data types: {df.dtypes}")
    if content_df is not None:
        st.write("Scraped Content DataFrame:")
        st.write(f"Shape: {content_df.shape}")
        st.write(f"Columns: {content_df.columns.tolist()}")
        st.write(f"Data types: {content_df.dtypes}")

def display_scraped_content(content_df):
    st.subheader("Scraped Content Overview")
    st.write(f"Number of records: {len(content_df)}")
    st.write(f"Columns: {', '.join(content_df.columns)}")

    numeric_cols = content_df.select_dtypes(include=[np.number]).columns
    categorical_cols = content_df.select_dtypes(include=['object']).columns

    if len(numeric_cols) > 0:
        st.subheader("Numeric Data Visualization")
        selected_numeric = st.selectbox("Choose a numeric column", numeric_cols)
        fig = px.histogram(content_df, x=selected_numeric, title=f"Distribution of {selected_numeric}")
        st.plotly_chart(fig, use_container_width=True)

        if len(numeric_cols) > 1:
            st.subheader("Scatter Plot")
            x_col = st.selectbox("Choose X axis", numeric_cols, key="x_axis")
            y_col = st.selectbox("Choose Y axis", [col for col in numeric_cols if col != x_col], key="y_axis")
            fig = px.scatter(content_df, x=x_col, y=y_col, title=f"{y_col} vs {x_col}")
            st.plotly_chart(fig, use_container_width=True)

    if len(categorical_cols) > 0:
        st.subheader("Categorical Data Visualization")
        selected_categorical = st.selectbox("Choose a categorical column", categorical_cols)
        fig = px.bar(content_df[selected_categorical].value_counts(), title=f"Counts of {selected_categorical}")
        st.plotly_chart(fig, use_container_width=True)


def display_table_view(df, content_df):
    st.subheader("AI Metrics")
    st.dataframe(df)
    if content_df is not None:
        st.subheader("Scraped Content")
        st.dataframe(content_df)


def display_download_options(df, content_df):
    col1, col2 = st.columns(2)
    with col1:
        csv = df.to_csv(index=False)
        st.download_button(
            label="Download AI Metrics as CSV",
            data=csv,
            file_name="ai_metrics.csv",
            mime="text/csv",
        )
    with col2:
        # Create Excel file using openpyxl
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "AI Metrics"

        # Write AI Metrics
        for row in dataframe_to_rows(df, index=False, header=True):
            ws1.append(row)

        if content_df is not None:
            ws2 = wb.create_sheet(title="Scraped Content")
            # Write Scraped Content
            for row in dataframe_to_rows(content_df, index=False, header=True):
                ws2.append(row)

        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # Create download link
        b64 = base64.b64encode(excel_buffer.getvalue()).decode()
        href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="scraped_data.xlsx">Download Excel file</a>'
        st.markdown(href, unsafe_allow_html=True)


def display_debug_info(df, content_df):
    st.write("Debug Info:")
    st.write("AI Metrics DataFrame:")
    st.write(f"Shape: {df.shape}")
    st.write(f"Columns: {df.columns.tolist()}")
    st.write(f"Data types: {df.dtypes}")
    if content_df is not None:
        st.write("Scraped Content DataFrame:")
        st.write(f"Shape: {content_df.shape}")
        st.write(f"Columns: {content_df.columns.tolist()}")
        st.write(f"Data types: {content_df.dtypes}")

def format_parsed_result(parsed_result: List[Dict[str, Any]], format_type: str = 'txt') -> str:
    if format_type == 'json':
        return json.dumps(parsed_result, indent=2)
    else:  # 'txt' format
        formatted = []
        for item in parsed_result:
            if 'error' in item:
                formatted.append(f"Error: {item['error']}")
            elif 'data' in item and 'content' in item['data']:
                content = item['data']['content']
                if isinstance(content, dict):
                    formatted.extend([f"{k}: {v}" for k, v in content.items()])
                elif isinstance(content, list):
                    formatted.extend([str(v) for v in content])
                else:
                    formatted.append(str(content))
            else:
                formatted.append(str(item))
        return "\n".join(formatted)


def get_preview(content: str, max_lines: int = 5) -> str:
    lines = content.split('\n')
    preview = '\n'.join(lines[:max_lines])
    if len(lines) > max_lines:
        preview += '\n...'
    return preview
